#!/usr/bin/env python3
"""
json_to_excel_and_csv.py

Pilih file JSON lewat dialog, lalu konversi menjadi:
 - 1 file Excel (.xlsx) dengan multi-sheet (sheet per table), dan
 - 1 file CSV gabungan (semua sheet dikombinasi). Jika pandas tersedia,
   CSV akan berisi kolom tambahan "__sheet" yang memberi tahu asal baris.
"""

import json
import os
import sys
from datetime import datetime

# GUI for file dialogs
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None

# try pandas
USE_PANDAS = False
try:
    import pandas as pd
    USE_PANDAS = True
except Exception:
    USE_PANDAS = False

# try openpyxl fallback
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

# ---------- utilities ----------
def print_instructions():
    print("=" * 60)
    print(" 📌 INSTRUKSI PENGGUNAAN PROGRAM JSON → Excel & CSV ")
    print("=" * 60)
    print("1. Pastikan sudah install dependency (sekali saja):")
    print("   pip install pandas openpyxl")
    print("   (jika tidak bisa install pandas, setidaknya: pip install openpyxl)\n")
    print("2. Setelah program berjalan, pilih file JSON yang ingin Anda konversi.")
    print("3. Tentukan nama & lokasi penyimpanan file Excel (.xlsx).")
    print("   Program juga akan membuat file CSV (.csv) dengan nama yang sama (hanya ekstensi berubah).")
    print("4. Selesai! Periksa folder yang Anda pilih untuk menemukan kedua file.\n")
    print("=" * 60, "\n")

def flatten(d, parent_key='', sep='.'):
    items = {}
    if isinstance(d, dict):
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.update(flatten(v, new_key, sep=sep))
            elif isinstance(v, list):
                items[new_key] = json.dumps(v, ensure_ascii=False)
            else:
                items[new_key] = v
    else:
        items[parent_key] = d
    return items

def make_sheets_from_obj(obj):
    sheets = {}
    if isinstance(obj, list):
        rows = []
        for item in obj:
            if isinstance(item, dict):
                rows.append(flatten(item))
            else:
                rows.append({'value': item})
        sheets['Sheet1'] = rows
        return sheets

    if isinstance(obj, dict):
        remaining = {}
        for k, v in obj.items():
            if isinstance(v, list):
                if all(isinstance(i, dict) for i in v):
                    rows = [flatten(i) for i in v]
                    sheets[str(k)] = rows
                else:
                    rows = [{'value': json.dumps(i, ensure_ascii=False)} for i in v]
                    sheets[str(k)] = rows
            else:
                remaining[k] = v
        if remaining:
            sheets['Summary'] = [flatten(remaining)]
        if not sheets:
            sheets['Sheet1'] = [flatten(obj)]
        return sheets

    sheets['Sheet1'] = [{'value': json.dumps(obj, ensure_ascii=False)}]
    return sheets

def sanitize_sheet_name(name, default='Sheet'):
    invalid = ['\\', '/', '*', '?', ':', '[', ']']
    s = ''.join('_' if ch in invalid else ch for ch in str(name))
    s = s.strip()
    if not s:
        s = default
    return s[:31]

# ---------- Excel writers ----------
def write_excel_with_pandas(sheets, filename):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for raw_name, rows in sheets.items():
            name = sanitize_sheet_name(raw_name, default='Sheet')
            if not rows:
                pd.DataFrame().to_excel(writer, sheet_name=name, index=False)
                continue
            try:
                df = pd.json_normalize(rows)
            except Exception:
                df = pd.DataFrame(rows)
            base = name
            i = 1
            while name in writer.sheets:
                name = (base[:28] + f"_{i}") if len(base) > 28 else f"{base}_{i}"
                i += 1
            df.to_excel(writer, sheet_name=name, index=False)

def write_excel_with_openpyxl(sheets, filename):
    if Workbook is None:
        raise RuntimeError("openpyxl tidak tersedia. Install: pip install openpyxl")
    wb = Workbook()
    default_sheet = wb.active
    first = True
    for raw_name, rows in sheets.items():
        name = sanitize_sheet_name(raw_name, default='Sheet')
        base = name
        i = 1
        while name in wb.sheetnames:
            name = (base[:28] + f"_{i}") if len(base) > 28 else f"{base}_{i}"
            i += 1
        if first:
            ws = default_sheet
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)
        if not rows:
            continue
        # headers collect
        headers = []
        for r in rows:
            for k in r.keys():
                if k not in headers:
                    headers.append(k)
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        for r_idx, r in enumerate(rows, start=2):
            for c_idx, h in enumerate(headers, start=1):
                val = r.get(h, "")
                ws.cell(row=r_idx, column=c_idx, value=val)
        # adjust widths
        for i_col, h in enumerate(headers, start=1):
            col_letter = get_column_letter(i_col)
            max_len = len(str(h))
            for cell in ws[col_letter]:
                if cell.value is not None:
                    l = len(str(cell.value))
                    if l > max_len:
                        max_len = l
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    wb.save(filename)

# ---------- CSV writers ----------
def write_combined_csv_with_pandas(sheets, csv_path):
    # sheets: dict name -> list of dict rows
    dfs = []
    for name, rows in sheets.items():
        if not rows:
            continue
        try:
            df = pd.json_normalize(rows)
        except Exception:
            df = pd.DataFrame(rows)
        df['__sheet'] = name
        dfs.append(df)
    if dfs:
        combined = pd.concat(dfs, ignore_index=True, sort=False)
        # place __sheet as first column
        cols = combined.columns.tolist()
        cols = ['__sheet'] + [c for c in cols if c != '__sheet']
        combined.to_csv(csv_path, index=False, columns=cols)
    else:
        # write empty csv
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write('')

def write_combined_csv_fallback(sheets, csv_path):
    # human-readable fallback: write sheet header then table rows, separate by blank line
    import csv
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        first_sheet = True
        for name, rows in sheets.items():
            if not first_sheet:
                writer.writerow([])  # blank line between sheets
            first_sheet = False
            # write a sheet label row
            writer.writerow([f"Sheet: {name}"])
            if not rows:
                continue
            # gather headers
            headers = []
            for r in rows:
                for k in r.keys():
                    if k not in headers:
                        headers.append(k)
            # write header
            writer.writerow(['__sheet'] + headers)
            # write rows
            for r in rows:
                row = [name] + [r.get(h, "") for h in headers]
                writer.writerow(row)

# ---------- File dialogs ----------
def pick_file_dialog():
    if tk is None:
        raise RuntimeError("tkinter tidak tersedia di environment ini.")
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    file_path = filedialog.askopenfilename(
        title="Pilih file JSON",
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
    )
    root.destroy()
    return file_path

def save_file_dialog(default_name):
    if tk is None:
        raise RuntimeError("tkinter tidak tersedia di environment ini.")
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    save_path = filedialog.asksaveasfilename(
        title="Simpan file Excel sebagai (program juga akan buat CSV dengan nama yang sama)",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    root.destroy()
    return save_path

# ---------- Main ----------
def main():
    print_instructions()

    # pick json file (or from argv)
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
        if not os.path.isfile(json_path):
            print("File tidak ditemukan:", json_path)
            return
    else:
        try:
            json_path = pick_file_dialog()
        except Exception as e:
            print("Gagal membuat dialog file. Error:", e)
            print("Sebagai alternatif, jalankan: python json_to_excel_and_csv.py path/to/file.json")
            return

    if not json_path:
        print("Tidak ada file dipilih. Keluar.")
        return

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            obj = json.load(f)
    except Exception as e:
        print("Gagal membaca/parse JSON:", e)
        return

    sheets = make_sheets_from_obj(obj)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = f"json_to_excel_{ts}.xlsx"

    try:
        save_xlsx_path = save_file_dialog(default_name)
    except Exception as e:
        print("Gagal membuat dialog save. Error:", e)
        save_xlsx_path = os.path.join(os.getcwd(), default_name)
        print("File akan disimpan di:", save_xlsx_path)

    if not save_xlsx_path:
        print("Penyimpanan dibatalkan. Keluar.")
        return

    # derive csv path (same folder, same base name, .csv)
    base, _ext = os.path.splitext(save_xlsx_path)
    csv_path = base + ".csv"

    try:
        # write xlsx
        if USE_PANDAS:
            write_excel_with_pandas(sheets, save_xlsx_path)
        else:
            write_excel_with_openpyxl(sheets, save_xlsx_path)

        # write combined csv
        if USE_PANDAS:
            write_combined_csv_with_pandas(sheets, csv_path)
        else:
            write_combined_csv_fallback(sheets, csv_path)

    except Exception as e:
        print("Gagal menyimpan file:", e)
        print("Pastikan dependencies telah diinstal: pip install pandas openpyxl")
        return

    print("\n✅ Sukses!")
    print("File Excel (.xlsx):", os.path.abspath(save_xlsx_path))
    print("File CSV (.csv)  : ", os.path.abspath(csv_path))
    print("\nSheet yang dibuat di Excel:")
    for name in sheets.keys():
        print("-", sanitize_sheet_name(name))

    # optional GUI notice
    if tk is not None:
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo("Selesai",
                f"File Excel:\n{os.path.abspath(save_xlsx_path)}\n\nFile CSV:\n{os.path.abspath(csv_path)}")
            root.destroy()
        except Exception:
            pass

if __name__ == "__main__":
    main()